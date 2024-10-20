from openpyxl import Workbook, load_workbook
import requests
from datetime import datetime, timedelta
import time
import math
import os 
from dotenv import load_dotenv
import os 
import json

# Inits data to a workbook to become usable
def init_data_wb(input_file):
    wb = load_workbook(input_file)
    new_wb = Workbook()
    return wb, new_wb

# Converts excel data to dict
def convert_to_dict(sheet):
    assignment_table = {0:"collection_period",1:"country",2:"site_id",3:"latitude",4:"longitude",5:"location",6:"town",7:"description",8:"species",9:"date",10:"set_up_time",11:"collection_time",12:"lifestage",13:"comments"}
    i = 0
    data = []
    period_info = {}
    j = 0
    for row in sheet.iter_rows(values_only=True):
        for val in row:
            period_info[assignment_table[i]] = val
            i += 1
            if i == 14:
                i = 0
                data.append(period_info.copy())
                period_info.clear()
    return data

# Used to insert columns, not used but useful for manipulation
def insert_column_at_index(sheet, new_column_data, insert_at):
    # Insert one blank column at the insert position
    sheet.insert_cols(insert_at)
    
    # Populate the new column with data (row by row)
    for row_idx, value in enumerate(new_column_data, start=1):
        sheet.cell(row=row_idx, column=insert_at, value=value) 

# gets daily historical weather : returns averages
def get_historical_weather(lat, lon, start_date, end_date):
    # Rate Limit of 3 requests per second on API
    time.sleep(.35)
    current_dir = os.getcwd()
    env_path = os.path.join(current_dir, '.env')
    load_dotenv(env_path)
    api_key = os.getenv("API_KEY")
    '''
    lat (string) (float)
    lon (string) ) (float)
    start_date (string) 2024-05-16
    end_date (string)
    '''
    
    url = "https://meteostat.p.rapidapi.com/point/daily"

    querystring = {"lat":lat,"lon":lon,"alt":"43","start":start_date,"end":end_date}

    headers = {
        "x-rapidapi-key": api_key,
        "x-rapidapi-host": "meteostat.p.rapidapi.com"
    }
    response = requests.get(url, headers=headers, params=querystring)
    data = response.json()['data']
    return data

# gets hourly historical weather
def get_historical_weather_hr(lat, lon, start_date, end_date):
    # Rate Limit of 3 requests per second on API
    time.sleep(.35)
    current_dir = os.getcwd()
    env_path = os.path.join(current_dir, '.env')
    load_dotenv(env_path)
    api_key = os.getenv("API_KEY")
    '''
    lat (string) (float)
    lon (string) ) (float)
    start_date (string) 2024-05-16
    end_date (string)
    '''
    
    url = "https://meteostat.p.rapidapi.com/point/hourly"

    querystring = {"lat":lat,"lon":lon,"alt":"43","start":start_date,"end":end_date}

    headers = {
        "x-rapidapi-key": api_key,
        "x-rapidapi-host": "meteostat.p.rapidapi.com"
    }
    response = requests.get(url, headers=headers, params=querystring)
    data = response.json()['data']
    return data

def create_data_xlsx(data,output_file,rain_threshold):
    i = 0   
    rows = []
    current_dir = os.getcwd()
    env_path = os.path.join(current_dir, '.env')
    load_dotenv(env_path)
    # Access data from excel data sheet
    for info_row in data:
        # column name row       
        if i == 0:
            i += 1 
            continue      
        row = [] 
        i+=1
        collection_date = info_row.get("date")
        if collection_date == None:
            continue      
        start_date = str(collection_date - timedelta(days=7)).split(" ")[0]
        end_date = str(collection_date + timedelta(days=7)).split(" ")[0]
        latitude = info_row.get("latitude")
        longitude = info_row.get("longitude")

        set_up_time = info_row.get("set_up_time").replace(minute=0, second=0, microsecond=0)
        collection_time = info_row.get("collection_time").replace(minute=0, second=0, microsecond=0)     
        trap_duration_time = datetime.combine(datetime.today(),collection_time) - datetime.combine(datetime.today() - timedelta(days=1),set_up_time)
        trap_hours = math.ceil(trap_duration_time.total_seconds() / 3600)
        
        daily_data = get_historical_weather(latitude,longitude,start_date,end_date)
        hourly_data = get_historical_weather_hr(latitude,longitude,start_date,end_date)
        
        # break total hourly data for start and end (+-7) into trap day
        # add avg relative humdity (rhum) to each daily
        trap_duration = []
        period = []
        total_rhum = 0
        i = 0
        prev_date = str(hourly_data[0].get("time")).split(" ")[0]
        hours_elapsed = 0
        trap_set = False
        for time in hourly_data:           
            curr_time = time.get("time")
            curr_date = str(curr_time).split(" ")[0]
            if curr_date != prev_date:
                daily_data[i]["avg_rhum"] = total_rhum / 24
                total_rhum = 0
                i += 1
            if curr_date == str(collection_date).split(" ")[0] and datetime.strptime(curr_time, "%Y-%m-%d %H:%M:%S").time() >= set_up_time:
                trap_set = True
            if trap_set and hours_elapsed <= trap_hours:
                only_time = str(curr_time).split(" ")[1]
                period.append(only_time)
                # Date
                period.append(curr_date)
                # Site ID 
                period.append(info_row.get("site_id"))
                # Latitude
                period.append(info_row.get("latitude"))
                # Longitude
                period.append(info_row.get("longitude"))
                # trap day
                period.append("Yes")
                # Percipitation
                period.append(time.get("prcp"))
                # Average temp
                period.append(time.get("temp"))
                # Min temp
                period.append(time.get("temp"))
                # max temp
                period.append(time.get("temp"))
                # humidity
                period.append(time.get("rhum"))
                # trap duration
                period.append(trap_duration_time)

                trap_duration.append(period)
                period = []
                hours_elapsed += 1
            if time.get("rhum"):
                total_rhum += time.get("rhum")
            prev_date = curr_date
        daily_data[i]["avg_rhum"] = total_rhum / 24

        # put into row format
        days_since_last_rain = "NA"
        for day in daily_data:
            row = []
            row.append("NA")
            # Date
            row.append(day.get("date"))
            # Site ID 
            row.append(info_row.get("site_id"))
            # Latitude
            row.append(info_row.get("latitude"))
            # Longitude
            row.append(info_row.get("longitude"))
            # Trap Day
            row.append("No")
            # Percipitation
            row.append(day.get("prcp"))
            # Average temp
            row.append(day.get("tavg"))
            # Min temp
            row.append(day.get("tmin"))
            # max temp
            row.append(day.get("tmax"))
            # humidity
            row.append(day.get("avg_rhum"))
            # trap duration
            row.append(trap_duration_time)
            # days since last rain
            row.append(days_since_last_rain)

            prcp = day.get("prcp")
            if prcp != None and prcp >= rain_threshold:
                days_since_last_rain = 0
            if day.get("date") == str(collection_date).split(" ")[0]:
                row[5] = "Yes"
                rows.append(row)
                row = []           
                for hour in trap_duration:
                    hour.append(days_since_last_rain)
                    rows.append(hour)
            else:
                rows.append(row)
            if days_since_last_rain != "NA":
                days_since_last_rain += 1  

    column_names = ["Time","Date","Site ID","Latitude","Longitude","Trap Day","Precipitation","Average Temperature", "Min Temperatures","Max Temperature","Humidity","Trap Duration","Days Since Last Rainfall"]
    wb = Workbook()
    ws = wb.active
    ws.append(column_names)
    for row in rows:
        ws.append(row) 
    wb.save(output_file)

def get_settings():
    current_dir = os.getcwd()
    settings_path = os.path.join(current_dir, 'settings.json')
    with open(settings_path) as f:
        settings = json.load(f)
    input_file = settings.get("input_file")
    output_file = settings.get("output_file")
    rain_threshold = float(settings.get("rain_threshold"))
    return input_file, output_file, rain_threshold

def main():
    input_file, output_file, rain_threshold = get_settings()
    wb, new_wb = init_data_wb(input_file)
    sheet = wb.active
    data = convert_to_dict(sheet)
    create_data_xlsx(data,output_file,rain_threshold)

if __name__=="__main__":
    main()